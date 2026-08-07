#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
商标商品清单生成脚本（trademark-assistant）
================================================
基于 templates/导入商品信息.xlsx 模板，把一份商品清单数据写成可导入
中国商标网申请系统的 Excel 文件。

输入（JSON 数组，每项含：类别、类似群、商品名称；序号自动生成）：

    [
      {"类别": 9, "类似群": "0901", "商品名称": "计算机软件（已录制）"},
      {"类别": 9, "类似群": "0907", "商品名称": "智能手机"}
    ]

用法：

    uv run --with openpyxl python scripts/script.py --input goods.json --output 星火-第9类-商品清单.xlsx
    echo '[{"类别":9,"类似群":"0901","商品名称":"计算机软件（已录制）"}]' | uv run --with openpyxl python scripts/script.py --output out.xlsx

安全规则：
    1. 目标文件已存在时默认拒绝覆盖，需显式 --force。
    2. 启动时校验模板表头，表头不符立即报错退出。
    3. 校验类别（1-45）、类似群（4 位数字）、商品名称非空。
"""

import argparse
import json
import re
import sys
from pathlib import Path

SKILL_DIR = Path(__file__).resolve().parent.parent
TEMPLATE = SKILL_DIR / "templates" / "导入商品信息.xlsx"
EXPECTED_HEADER = ["序号", "商品类别", "类似群", "商品名称"]


def log(msg: str) -> None:
    print(msg, file=sys.stderr)


def require_openpyxl():
    """按需加载硬依赖，使纯数据校验可以在无 openpyxl 环境中测试。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        log(
            "[错误] 未安装 openpyxl。请用以下方式运行：\n"
            "    uv run --with openpyxl python scripts/script.py ...\n"
            "或先执行： python3 -m pip install openpyxl"
        )
        raise SystemExit(2)
    return load_workbook


def parse_category(idx: int, value):
    """只接受真正的整数或纯数字字符串，拒绝 9.5、9.0、True 等隐式转换。"""
    if isinstance(value, bool):
        return None, f"第 {idx} 项 类别={value!r} 不是 1-45 的整数"
    if isinstance(value, int):
        category = value
    elif isinstance(value, str) and re.fullmatch(r"\d+", value.strip()):
        category = int(value.strip())
    else:
        return None, f"第 {idx} 项 类别={value!r} 不是 1-45 的整数"

    if not 1 <= category <= 45:
        return None, f"第 {idx} 项 类别={value!r} 超出 1-45 范围"
    return category, None


def normalize_item(idx: int, item):
    errs = []
    if not isinstance(item, dict):
        return None, [f"第 {idx} 项不是 JSON 对象"]

    cat_int, cat_error = parse_category(idx, item.get("类别"))
    if cat_error:
        errs.append(cat_error)

    group_value = item.get("类似群")
    if isinstance(group_value, str):
        grp = group_value.strip()
    else:
        grp = ""
        errs.append(f"第 {idx} 项 类似群必须是 4 位数字字符串，例如 '0901'")

    if not (len(grp) == 4 and grp.isdigit()):
        if isinstance(group_value, str):
            errs.append(f"第 {idx} 项 类似群={grp!r} 不是 4 位数字")
    elif cat_int is not None and int(grp[:2]) != cat_int:
        errs.append(f"第 {idx} 项 类似群 {grp} 与类别 {cat_int} 不一致")

    name_value = item.get("商品名称")
    if not isinstance(name_value, str):
        name = ""
        errs.append(f"第 {idx} 项 商品名称必须是字符串")
    else:
        name = name_value.strip()
        if not name:
            errs.append(f"第 {idx} 项 商品名称为空")

    if errs:
        return None, errs
    return {"类别": cat_int, "类似群": grp, "商品名称": name}, []


def load_items(args) -> list:
    if args.input and args.input != "-":
        raw = Path(args.input).read_text(encoding="utf-8")
    else:
        raw = sys.stdin.read()

    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        log(f"[错误] 输入 JSON 解析失败：{e}")
        sys.exit(1)

    if not isinstance(data, list) or not data:
        log("[错误] 输入必须是商品数组（至少 1 项）。")
        sys.exit(1)

    errs = []
    normalized = []
    for i, it in enumerate(data, start=1):
        item, item_errs = normalize_item(i, it)
        errs.extend(item_errs)
        if item is not None:
            normalized.append(item)
    if errs:
        log("[错误] 数据校验未通过：")
        for e in errs:
            log("  - " + e)
        sys.exit(1)

    return normalized


def main() -> None:
    ap = argparse.ArgumentParser(description="生成可导入商标系统的商品清单 Excel")
    ap.add_argument("--input", "-i", default="-", help="商品清单 JSON 路径，- 或缺省读 stdin")
    ap.add_argument("--output", "-o", required=True, help="输出 Excel 路径")
    ap.add_argument("--force", action="store_true", help="目标已存在时允许覆盖")
    args = ap.parse_args()

    if not TEMPLATE.exists():
        log(f"[错误] 找不到模板文件：{TEMPLATE}")
        sys.exit(1)

    out = Path(args.output).resolve()
    if out.exists() and not args.force:
        log(f"[错误] 输出文件已存在：{out}\n  如确认覆盖，请加 --force。")
        sys.exit(1)

    items = load_items(args)

    load_workbook = require_openpyxl()
    wb = load_workbook(TEMPLATE)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if header != EXPECTED_HEADER:
        log(f"[错误] 模板表头与预期不符。\n  预期：{EXPECTED_HEADER}\n  实际：{header}")
        sys.exit(1)

    # 清空模板中第 2 行起的残留示例
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for i, it in enumerate(items, start=1):
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=it["类别"])
        ws.cell(row=i + 1, column=3, value=it["类似群"])
        ws.cell(row=i + 1, column=4, value=it["商品名称"])

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    log(f"[完成] 已生成 {len(items)} 条商品 → {out}")


if __name__ == "__main__":
    main()
