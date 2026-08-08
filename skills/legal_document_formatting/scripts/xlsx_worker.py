#!/usr/bin/env python3
"""Compact deterministic XLSX worker for LegalWork.

The worker deliberately supports a small set of high-value spreadsheet actions
and returns compact JSON only. It does not dump workbook XML or full cell data
into model history.
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
except Exception as exc:  # pragma: no cover
    Workbook = None  # type: ignore[assignment]
    IMPORT_ERROR = str(exc)
else:
    IMPORT_ERROR = ""

MAX_SHEETS = 64
MAX_PREVIEW_ROWS = 8
MAX_PREVIEW_COLS = 8


def emit(payload: dict[str, Any], code: int = 0) -> None:
    print(json.dumps(payload, ensure_ascii=False, separators=(",", ":")))
    raise SystemExit(code)


def require_openpyxl(operation: str) -> None:
    if Workbook is None:
        emit({"status": "error", "operation": operation, "error": f"openpyxl unavailable: {IMPORT_ERROR}"}, 1)


def workbook_path(raw: str, operation: str) -> Path:
    path = Path(raw).expanduser().resolve()
    if not path.is_file():
        emit({"status": "error", "operation": operation, "error": f"file not found: {path}"}, 1)
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        emit({"status": "error", "operation": operation, "error": "xlsx worker supports .xlsx/.xlsm only"}, 1)
    return path


def output_path(raw: str, operation: str) -> Path:
    path = Path(raw).expanduser().resolve()
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        emit({"status": "error", "operation": operation, "error": "output must end with .xlsx or .xlsm"}, 1)
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def preview_sheet(sheet: Any) -> list[list[Any]]:
    rows: list[list[Any]] = []
    max_row = min(sheet.max_row or 0, MAX_PREVIEW_ROWS)
    max_col = min(sheet.max_column or 0, MAX_PREVIEW_COLS)
    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
        rows.append([value for value in row])
    return rows


def cmd_inspect(args: argparse.Namespace) -> None:
    operation = "xlsx-inspect"
    require_openpyxl(operation)
    path = workbook_path(args.input, operation)
    keep_vba = path.suffix.lower() == ".xlsm"
    book = load_workbook(str(path), read_only=False, data_only=False, keep_vba=keep_vba)
    sheets = []
    for sheet in book.worksheets[:MAX_SHEETS]:
        sheets.append({
            "name": sheet.title,
            "rows": sheet.max_row,
            "columns": sheet.max_column,
            "merged_ranges": len(sheet.merged_cells.ranges),
            "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
            "preview": preview_sheet(sheet) if args.preview else None,
        })
    emit({
        "status": "ok",
        "operation": operation,
        "input": str(path),
        "sheet_count": len(book.sheetnames),
        "sheets": sheets,
        "truncated": len(book.sheetnames) > MAX_SHEETS,
    })


def cmd_from_json(args: argparse.Namespace) -> None:
    operation = "xlsx-from-json"
    require_openpyxl(operation)
    spec_path = Path(args.spec).expanduser().resolve()
    if not spec_path.is_file():
        emit({"status": "error", "operation": operation, "error": f"spec not found: {spec_path}"}, 1)
    try:
        payload = json.loads(spec_path.read_text(encoding="utf-8"))
    except Exception as exc:
        emit({"status": "error", "operation": operation, "error": f"invalid JSON spec: {exc}"}, 1)
    sheets = payload.get("sheets") if isinstance(payload, dict) else None
    if not isinstance(sheets, list) or not sheets:
        emit({"status": "error", "operation": operation, "error": "spec must contain a non-empty sheets array"}, 1)
    output = output_path(args.output, operation)
    book = Workbook()
    book.remove(book.active)
    written_cells = 0
    for index, sheet_spec in enumerate(sheets[:MAX_SHEETS]):
        if not isinstance(sheet_spec, dict):
            continue
        title = str(sheet_spec.get("name") or f"Sheet{index + 1}")[:31]
        sheet = book.create_sheet(title=title)
        rows = sheet_spec.get("rows")
        if isinstance(rows, list):
            for row in rows:
                if not isinstance(row, list):
                    continue
                sheet.append(row)
                written_cells += len(row)
        widths = sheet_spec.get("column_widths")
        if isinstance(widths, dict):
            for column, width in widths.items():
                try:
                    sheet.column_dimensions[str(column)].width = float(width)
                except Exception:
                    continue
        freeze = sheet_spec.get("freeze_panes")
        if isinstance(freeze, str) and freeze.strip():
            sheet.freeze_panes = freeze.strip()
        autofilter = sheet_spec.get("auto_filter")
        if isinstance(autofilter, str) and autofilter.strip():
            sheet.auto_filter.ref = autofilter.strip()
    book.save(str(output))
    emit({
        "status": "ok",
        "operation": operation,
        "output": str(output),
        "sheet_count": len(book.sheetnames),
        "written_cells": written_cells,
    })


def cmd_replace(args: argparse.Namespace) -> None:
    operation = "xlsx-replace"
    require_openpyxl(operation)
    source = workbook_path(args.input, operation)
    output = output_path(args.output, operation)
    keep_vba = source.suffix.lower() == ".xlsm" or output.suffix.lower() == ".xlsm"
    book = load_workbook(str(source), read_only=False, data_only=False, keep_vba=keep_vba)
    replaced = 0
    for sheet in book.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or args.find not in cell.value:
                    continue
                cell.value = cell.value.replace(args.find, args.replace)
                replaced += 1
    book.save(str(output))
    emit({"status": "ok", "operation": operation, "output": str(output), "replaced_cells": replaced})


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="LegalWork deterministic XLSX worker")
    sub = parser.add_subparsers(dest="command", required=True)

    inspect = sub.add_parser("inspect")
    inspect.add_argument("--input", required=True)
    inspect.add_argument("--preview", action="store_true")
    inspect.set_defaults(func=cmd_inspect)

    create = sub.add_parser("from-json")
    create.add_argument("--spec", required=True)
    create.add_argument("--output", required=True)
    create.set_defaults(func=cmd_from_json)

    replace = sub.add_parser("replace")
    replace.add_argument("--input", required=True)
    replace.add_argument("--output", required=True)
    replace.add_argument("--find", required=True)
    replace.add_argument("--replace", required=True)
    replace.set_defaults(func=cmd_replace)
    return parser


def main() -> None:
    args = build_parser().parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
